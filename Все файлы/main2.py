# -*- coding: UTF-8 -*-
import sys
import random

from openpyxl import Workbook, load_workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

# import xlsxwriter
# import xlwings as xw

reload(sys)
sys.setdefaultencoding('utf8')

# wb2 = load_workbook('labWork10.xlsx')
# ws = wb2['Фильтр']

# print ws._charts
# print wb2.get_sheet_names()
# print type(ws.auto_filter.ref)
# print ws.auto_filter
# first_cell = ws.auto_filter.ref.split(':')[0]
# cell_date_filter = ''
# cell_cost_filter = ''
# # print ws[first_cell].row
# for Colfilter in ws.auto_filter.filterColumn:
#     if Colfilter.filters is not None:
#         cell_date_filter = ws[first_cell].column + str(Colfilter.colId)
#         print type(cell_date_filter)
#         for Colfilter2 in Colfilter.filters.dateGroupItem:
#             print Colfilter2.year
#     print Colfilter.colId #относительно ws.auto_filter.ref
#     if Colfilter.customFilters is not None:
#         # print Colfilter.colId
#         for Colfilter1 in Colfilter.customFilters.customFilter:
#             print Colfilter1.operator+": "+Colfilter1.val
#             pass


def set_border_and_fill(ws, cell_range, fill=None):
    border_side = Side(border_style='thin', color='000000')
    border = Border(left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side)

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill
    return ws

def set_table_header(ws, headers_names):
    ws['B2'] = 'Поступление товаров'
    ws.merge_cells('B2:G2')
    ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
    ws = set_border_and_fill(ws, 'B4:G4', fill=PatternFill("solid", fgColor="DDDDDD"))
    for i in range(1, len(headers_names)+1):
        ws.cell(row=4, column=i+1).value = headers_names[i-1]

    return ws

def create_template(ws, data):

    ws.column_dimensions["B"].width = 5.0
    ws.column_dimensions["C"].width = 21.0
    ws.column_dimensions["D"].width = 18.0
    ws.column_dimensions["E"].width = 12.0
    ws.column_dimensions["F"].width = 12.0
    ws.column_dimensions["G"].width = 14.0

    ws.row_dimensions[4].height = 27

    shuffle_data = list(data)
    random.shuffle(shuffle_data)
    headers_names = ["№", "Наименование товара", "Дата поступления", "Количество", "Цена", "Стоимость"]
    ws = set_table_header(ws, headers_names)

    for i in range(1, len(shuffle_data)+1):
        for j in range(1, len(headers_names)+1):
            pos_i = i + 4
            ws.cell(row=pos_i, column=j).alignment = Alignment(horizontal="center", vertical="center")
            if j == 2:
                ws.cell(row=pos_i, column=j).value = i
            if j == 3:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][0]
            if j == 4:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][1]
            if j == 5:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][2]
            if j == 6:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][3]
            if j == 7:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][4]

    set_border_and_fill(ws, 'B5:G19')
    print '\n'
    return ws

wb = Workbook()
ws = wb.active

data = [
    ["Комбайн", "19.07.2017", 100, 7800.00],
    ["Миксер", "30.05.2017", 38, 3000.00],
    ["Микровоновка", "23.08.2017", 38, 4500.00],
    ["Пылесос", "17.03.2017", 25, 3000.00],
    ["Холодильник", "03.05.2016", 56, 25000.00],
    ["Пылесос", "03.08.2017", 6, 1500.00],
    ["Телевизор", "02.03.2014", 50, 6000.00],
    ["Телевизор", "16.02.2016", 19, 12000.00],
    ["Телевизор", "13.09.2017", 32, 4500.00],
    ["Утюг", "12.07.2016", 70, 2000.00],
    ["Утюг", "20.08.2016", 15, 1000.00],
    ["Утюг", "02.08.2017", 20, 2900.00],
    ["Чайник", "15.03.2017", 25, 1540.00],
    ["Чайник", "27.07.2016", 102, 1200.00],
    ["Чайник", "04.08.2016", 45, 500.00],
]


ws = create_template(ws, data)

wb.save('lab3_template.xlsx')

print len(data)










