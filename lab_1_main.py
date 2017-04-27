# -*- coding: UTF-8 -*-
import sys
import random
import datetime
import time
import json

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import *
from lab_1_create_template import lab_1_create_template
from lab_1_check_answer import lab_1_check_answer

reload(sys)
sys.setdefaultencoding('utf8')

# Создание шаблона
template_wb = Workbook()
template_ws = template_wb.active
employees = ["Иванов И.М.", "Коробова П.Н", "Морозов И.Р.", "Петров Г.Т.", "Ромашова П.Т.", "Смирнов С.И.", "Соколова О.С."]

template_ws = lab_1_create_template(template_ws, employees)
template_wb.save('lab1_template.xlsx')


# Проверка
student_wb =  load_workbook('lab1_student.xlsx')
student_wb_data_only =  load_workbook('lab1_student.xlsx', data_only=True)

result = lab_1_check_answer(student_wb, student_wb_data_only, employees)

print result
