# -*- coding: UTF-8 -*-
import sys
import random

from openpyxl import Workbook, load_workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

# import xlsxwriter
# import xlwings as xw

reload(sys)
sys.setdefaultencoding('utf8')

def ws_have_rule(ws, cells_range, operator, formula_value):
    for rule in ws.conditional_formatting.cf_rules.items():
        if cells_range == rule[0]:
            if operator == rule[1][0].operator:
                if int(rule[1][0].formula[0]) == formula_value:
                    return True
    return False



wb2 = load_workbook('lab3_template.xlsx')
ws = wb2['Sheet']

ws_have_rule(ws, 'F5:F19', 'lessThan', 5000)



