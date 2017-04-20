# -*- coding: UTF-8 -*-
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout
# import xlsxwriter
# import xlwings as xw

reload(sys)
sys.setdefaultencoding('utf8')

# wb2 = load_workbook('lab1_template.xlsx')
# ws = wb2[wb2.get_sheet_names()[0]]

# print ws._charts
# print wb2.get_sheet_names()
# print type(ws.auto_filter.ref)
# # print ws.auto_filter
# first_cell = ws.auto_filter.ref.split(':')[0]
# cell_date_filter = ''
# cell_cost_filter = ''
# print ws[first_cell].row
# for Colfilter in ws.auto_filter.filterColumn:
#     if Colfilter.filters is not None:
#         cell_date_filter = ws[first_cell].column + str(Colfilter.colId)
#         print cell_date_filter
#         for Colfilter2 in Colfilter.filters.dateGroupItem:
#             print Colfilter2.year
#     print Colfilter.colId #относительно ws.auto_filter.ref
#     if Colfilter.customFilters is not None:
#         # print Colfilter.colId
#         for Colfilter1 in Colfilter.customFilters.customFilter:
#             print Colfilter1.operator+": "+Colfilter1.val
#             pass



def string_is_money_format(str):
    if '₽' in str:
        return True
    if '£' in str:
        return True
    if '$' in str:
        return True
    if '€' in str:
        return True
    return False


def range_is_money_format(range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not string_is_money_format(cell.number_format):
                return False
    return True

def range_values_array(ws, range):
    arr = []
    rows = ws[range]
    for row in rows:
        for cell in row:
            arr.append(cell.value)
    return arr

def range_values_array_numeric(ws, range):
    arr = []
    rows = ws[range]
    for row in rows:
        for cell in row:
            arr.append(float(cell.value))
    return arr

def print_range(range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            print cell.coordinate, cell.value


def calculate_correct_values(range, dollar_rate):
    data = {}
    rows = ws[range]
    values = []
    for row in rows:
        for cell in row:
            values.append(float(cell.value))
    data["salary"] = {"values": values}

    premium = []
    total = []
    income_tax = []
    amount_granted = []
    amount_granted_dollar = []

    for v in data["salary"]["values"]:
        premium_val = v * 0.2
        premium.append(premium_val)
        total_val = premium_val + v
        total.append(total_val)
        income_tax_val = round(total_val*0.13, 2)
        income_tax.append(income_tax_val)
        amount_granted_val = (total_val - income_tax_val)
        amount_granted.append(amount_granted_val)
        amount_granted_dollar_val = round(amount_granted_val/dollar_rate, 2)
        amount_granted_dollar.append(amount_granted_dollar_val)

    data["premium"] = {"values": premium}
    data["total"] = {"values": total}
    data["total_sum"] = sum(total)
    data["income_tax"] = {"values": income_tax}
    data["income_tax_sum"] = sum(income_tax)
    data["amount_granted"] = {"values": amount_granted}
    data["amount_granted_sum"] = sum(amount_granted)
    data["amount_granted_dollar"] = {"values": amount_granted_dollar}
    data["amount_granted_dollar_sum"] = sum(amount_granted_dollar)
    data["avg_salary"] = round(reduce(lambda x, y: x + y, amount_granted) / len(amount_granted), 2)
    data["min_salary"] = round(min(amount_granted), 2)
    data["max_salary"] = round(max(amount_granted), 2)
    return data


def get_students_formulas_values(ws, s, e, ws_data_only):
    data = {}
    data["names"] = range_values_array(ws, 'B'+str(s)+':'+'B'+str(e))

    salary =  range_values_array_numeric(ws, 'E'+str(s)+':'+'E'+str(e))
    data["salary"] = {"values": salary}

    premium =  range_values_array(ws, 'F'+str(s)+':'+'F'+str(e))
    data["premium"] = {}
    data["premium"]["formula"] = premium
    premium = range_values_array_numeric(ws_data_only, 'F' + str(s) + ':' + 'F' + str(e))
    data["premium"]["values"]= premium

    total =  range_values_array(ws, 'G'+str(s)+':'+'G'+str(e))
    data["total"] = {}
    data["total_sum"] = {}
    data["total"]["formula"] =  total
    total = range_values_array_numeric(ws_data_only, 'G' + str(s) + ':' + 'G' + str(e))
    data["total"]["values"] = total
    data["total_sum"]["values"] = round(float(ws_data_only['G' + str(e + 1)].value), 2)
    data["total_sum"]["formula"] = ws['G' + str(e + 1)].value

    income_tax =  range_values_array(ws, 'H'+str(s)+':'+'H'+str(e))
    data["income_tax"] = {}
    data["income_tax_sum"] = {}
    data["income_tax"]["formula"] =  income_tax
    income_tax = range_values_array_numeric(ws_data_only, 'H' + str(s) + ':' + 'H' + str(e))
    data["income_tax"]["values"] = income_tax
    data["income_tax_sum"]["values"] = round(float(ws_data_only['H' + str(e + 1)].value), 2)
    data["income_tax_sum"]["formula"] = ws['H' + str(e + 1)].value

    amount_granted =  range_values_array(ws, 'I'+str(s)+':'+'I'+str(e))
    data["amount_granted"] = {}
    data["amount_granted_sum"] = {}
    data["amount_granted"]["formula"] = amount_granted
    amount_granted = range_values_array_numeric(ws_data_only, 'I' + str(s) + ':' + 'I' + str(e))
    data["amount_granted"]["values"] = amount_granted
    data["amount_granted_sum"]["values"] = round(float(ws_data_only['I' + str(e + 1)].value), 2)
    data["amount_granted_sum"]["formula"] = ws['I' + str(e + 1)].value

    amount_granted_dollar =  range_values_array(ws, 'J'+str(s)+':'+'J'+str(e))
    data["amount_granted_dollar"] = {}
    data["amount_granted_dollar_sum"] = {}
    data["amount_granted_dollar"]["formula"] =  amount_granted_dollar
    amount_granted_dollar = range_values_array_numeric(ws_data_only, 'J' + str(s) + ':' + 'J' + str(e))
    data["amount_granted_dollar"]["values"] = amount_granted_dollar
    data["amount_granted_dollar_sum"]["values"] = round(float(ws_data_only['J' + str(e + 1)].value), 2)
    data["amount_granted_dollar_sum"]["formula"] = ws['J' + str(e + 1)].value

    data["avg_salary"] = {}
    data["max_salary"] = {}
    data["min_salary"] = {}

    data["avg_salary"]["value"] = round(float(ws_data_only['C'+ str(e + 4)].value), 2)
    data["max_salary"]["value"] = round(float(ws_data_only['C'+ str(e + 5)].value), 2)
    data["min_salary"]["value"] = round(float(ws_data_only['C'+ str(e + 6)].value), 2)

    data["avg_salary"]["formula"] = ws['C'+ str(e + 4)].value
    data["max_salary"]["formula"] = ws['C'+ str(e + 5)].value
    data["min_salary"]["formula"] = ws['C'+ str(e + 6)].value

    return data


wb2 = load_workbook('lab1_template.xlsx')
ws = wb2[wb2.get_sheet_names()[0]]

wb2_read_only = load_workbook('lab1_template.xlsx', data_only=True)
ws_read_only = wb2_read_only[wb2_read_only.get_sheet_names()[0]]


employees = ["Иванов И.М.", "Коробова П.Н", "Морозов И.Р.", "Ромашова П.Т.", "Петров Г.Т.", "Смирнов С.И.", "Соколова О.С."]

start_row = 5
end_row = start_row + len(employees) - 1
range_salary = 'E5:E'+ str(end_row)
dollar_rate = 48
# print range_salary

correct_values_data = {}
correct_values_data = calculate_correct_values(range_salary, dollar_rate)

student_values_data = {}
student_values_data = get_students_formulas_values(ws, start_row, end_row, ws_read_only)

print student_values_data






# print_range('F5:F11')

# print range_is_money_format('E5:F11')

# print "D7: ", ws['F3'].number_format
# print "D7: ", ws['F4'].number_format
# print "D7: ", ws['F5'].number_format
#
# print "D7: ", ws['F6'].number_format

#  data_only=True
