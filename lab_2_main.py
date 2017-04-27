# -*- coding: UTF-8 -*-
import sys
import random
import datetime

from openpyxl import Workbook, load_workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
from openpyxl.drawing.image import Image

from lab_2_create_template import lab_2_create_template
from lab_2_check_answer import lab_2_check_answer

reload(sys)
sys.setdefaultencoding('utf8')



template_wb = Workbook()

template_wb = lab_2_create_template(template_wb)
template_wb.save('lab2_template.xlsx')

correct_wb = load_workbook('lab2_correct.xlsx')
correct_wb_data_only =  load_workbook('lab2_correct.xlsx', data_only=True)

# Проверка
student_wb =  load_workbook('lab2_student.xlsx')
student_wb_data_only =  load_workbook('lab2_correct.xlsx', data_only=True)

result = lab_2_check_answer(correct_wb, correct_wb_data_only, student_wb, student_wb_data_only)

print result