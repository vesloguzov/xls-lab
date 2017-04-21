# -*- coding: UTF-8 -*-
import sys
import random
import datetime

from openpyxl import Workbook, load_workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

reload(sys)
sys.setdefaultencoding('utf8')

def set_border_and_fill(ws, cell_range, fill=None):
    border_side = Side(border_style='thin', color='000000')
    border = Border(left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side)

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill
    return ws

def set_table_header(ws, headers_names):
    ws['B2'] = 'Поступление товаров'
    ws.merge_cells('B2:G2')
    ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
    ws = set_border_and_fill(ws, 'B4:G4', fill=PatternFill("solid", fgColor="DDDDDD"))
    for i in range(1, len(headers_names)+1):
        ws.cell(row=4, column=i+1).value = headers_names[i-1]

    return ws

def create_template(ws, data):

    ws.column_dimensions["B"].width = 5.0
    ws.column_dimensions["C"].width = 21.0
    ws.column_dimensions["D"].width = 18.0
    ws.column_dimensions["E"].width = 12.0
    ws.column_dimensions["F"].width = 12.0
    ws.column_dimensions["G"].width = 14.0

    ws.row_dimensions[4].height = 27

    shuffle_data = list(data)
    random.shuffle(shuffle_data)
    headers_names = ["№", "Наименование товара", "Дата поступления", "Количество", "Цена", "Стоимость"]
    ws = set_table_header(ws, headers_names)

    for i in range(1, len(shuffle_data)+1):
        for j in range(1, len(headers_names)+1):
            pos_i = i + 4
            ws.cell(row=pos_i, column=j).alignment = Alignment(horizontal="center", vertical="center")
            if j == 2:
                ws.cell(row=pos_i, column=j).value = i
            if j == 3:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][0]
            if j == 4:

                ws.cell(row=pos_i, column=j).value = datetime.datetime.strptime(shuffle_data[i-1][1], "%d.%m.%Y")
                ws.cell(row=pos_i, column=j).number_format = 'MM/DD/YYYY'
            if j == 5:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][2]
            if j == 6:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][3]
            if j == 7:
                ws.cell(row=pos_i, column=j).value = shuffle_data[i-1][4]

    set_border_and_fill(ws, 'B5:G19')

    return ws
