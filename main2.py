# -*- coding: UTF-8 -*-
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout
# import xlsxwriter
# import xlwings as xw

reload(sys)
sys.setdefaultencoding('utf8')

wb2 = load_workbook('data/rabota 8.xlsx')
ws = wb2[wb2.get_sheet_names()[0]]


# print wb2.get_sheet_names()
# print type(ws.auto_filter.ref)
# # print ws.auto_filter
# first_cell = ws.auto_filter.ref.split(':')[0]
# cell_date_filter = ''
# cell_cost_filter = ''
# print ws[first_cell].row
# for Colfilter in ws.auto_filter.filterColumn:
#     if Colfilter.filters is not None:
#         cell_date_filter = ws[first_cell].column + str(Colfilter.colId)
#         print cell_date_filter
#         for Colfilter2 in Colfilter.filters.dateGroupItem:
#             print Colfilter2.year
#     print Colfilter.colId #относительно ws.auto_filter.ref
#     if Colfilter.customFilters is not None:
#         # print Colfilter.colId
#         for Colfilter1 in Colfilter.customFilters.customFilter:
#             print Colfilter1.operator+": "+Colfilter1.val
#             pass


# interesCell = 'G4'
#
# print "D7: ", ws[interesCell].value
# print "D7: ", ws[interesCell].coordinate
# print "D7: ", ws[interesCell].column
# print "D7: ", ws[interesCell].base_date
# print "D7: ", ws[interesCell].guess_types
# print "D7: ", ws[interesCell].internal_value
# print "D7: ", ws[interesCell].is_date
# print "D7: ", ws[interesCell].number_format


l = reader('''<?xml version="1.0" encoding="UTF-8"?>
<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:c16r2="http://schemas.microsoft.com/office/drawing/2015/06/chart" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
   <c:date1904 val="0" />
   <c:lang val="ru-RU" />
   <c:roundedCorners val="0" />
   <mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006">
      <mc:Choice xmlns:c14="http://schemas.microsoft.com/office/drawing/2007/8/2/chart" Requires="c14">
         <c14:style val="102" />
      </mc:Choice>
      <mc:Fallback>
         <c:style val="2" />
      </mc:Fallback>
   </mc:AlternateContent>
   <c:chart>
      <c:title>
         <c:tx>
            <c:rich>
               <a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1" />
               <a:lstStyle />
               <a:p>
                  <a:pPr>
                     <a:defRPr sz="1400" b="0" i="0" u="none" strike="noStrike" kern="1200" spc="0" baseline="0">
                        <a:solidFill>
                           <a:schemeClr val="tx1">
                              <a:lumMod val="65000" />
                              <a:lumOff val="35000" />
                           </a:schemeClr>
                        </a:solidFill>
                        <a:latin typeface="+mn-lt" />
                        <a:ea typeface="+mn-ea" />
                        <a:cs typeface="+mn-cs" />
                     </a:defRPr>
                  </a:pPr>
                  <a:r>
                     <a:rPr lang="ru-RU" />
                     <a:t>ЗП</a:t>
                  </a:r>
               </a:p>
            </c:rich>
         </c:tx>
         <c:layout />
         <c:overlay val="0" />
         <c:spPr>
            <a:noFill />
            <a:ln>
               <a:noFill />
            </a:ln>
            <a:effectLst />
         </c:spPr>
         <c:txPr>
            <a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1" />
            <a:lstStyle />
            <a:p>
               <a:pPr>
                  <a:defRPr sz="1400" b="0" i="0" u="none" strike="noStrike" kern="1200" spc="0" baseline="0">
                     <a:solidFill>
                        <a:schemeClr val="tx1">
                           <a:lumMod val="65000" />
                           <a:lumOff val="35000" />
                        </a:schemeClr>
                     </a:solidFill>
                     <a:latin typeface="+mn-lt" />
                     <a:ea typeface="+mn-ea" />
                     <a:cs typeface="+mn-cs" />
                  </a:defRPr>
               </a:pPr>
               <a:endParaRPr lang="ru-RU" />
            </a:p>
         </c:txPr>
      </c:title>
      <c:autoTitleDeleted val="0" />
      <c:plotArea>
         <c:layout />
         <c:barChart>
            <c:barDir val="col" />
            <c:grouping val="clustered" />
            <c:varyColors val="0" />
            <c:ser>
               <c:idx val="0" />
               <c:order val="0" />
               <c:spPr>
                  <a:solidFill>
                     <a:schemeClr val="accent1" />
                  </a:solidFill>
                  <a:ln>
                     <a:noFill />
                  </a:ln>
                  <a:effectLst />
               </c:spPr>
               <c:invertIfNegative val="0" />
               <c:cat>
                  <c:strRef>
                     <c:f>Sheet!$B$5:$B$11</c:f>
                     <c:strCache>
                        <c:ptCount val="7" />
                        <c:pt idx="0">
                           <c:v>Иванов И.М.</c:v>
                        </c:pt>
                        <c:pt idx="1">
                           <c:v>Коробова П.Н</c:v>
                        </c:pt>
                        <c:pt idx="2">
                           <c:v>Морозов И.Р.</c:v>
                        </c:pt>
                        <c:pt idx="3">
                           <c:v>Ромашова П.Т.</c:v>
                        </c:pt>
                        <c:pt idx="4">
                           <c:v>Петров Г.Т.</c:v>
                        </c:pt>
                        <c:pt idx="5">
                           <c:v>Смирнов С.И.</c:v>
                        </c:pt>
                        <c:pt idx="6">
                           <c:v>Соколова О.С.</c:v>
                        </c:pt>
                     </c:strCache>
                  </c:strRef>
               </c:cat>
               <c:val>
                  <c:numRef>
                     <c:f>Sheet!$I$5:$I$11</c:f>
                     <c:numCache>
                        <c:formatCode>#\ ##0.00\ "₽"</c:formatCode>
                        <c:ptCount val="7" />
                        <c:pt idx="0">
                           <c:v>1044</c:v>
                        </c:pt>
                        <c:pt idx="1">
                           <c:v>2088</c:v>
                        </c:pt>
                        <c:pt idx="2">
                           <c:v>3132</c:v>
                        </c:pt>
                        <c:pt idx="3">
                           <c:v>3654</c:v>
                        </c:pt>
                        <c:pt idx="4">
                           <c:v>5794.2</c:v>
                        </c:pt>
                        <c:pt idx="5">
                           <c:v>5220</c:v>
                        </c:pt>
                        <c:pt idx="6">
                           <c:v>83520</c:v>
                        </c:pt>
                     </c:numCache>
                  </c:numRef>
               </c:val>

            </c:ser>
            <c:dLbls>
               <c:showLegendKey val="0" />
               <c:showVal val="0" />
               <c:showCatName val="0" />
               <c:showSerName val="0" />
               <c:showPercent val="0" />
               <c:showBubbleSize val="0" />
            </c:dLbls>
            <c:gapWidth val="219" />
            <c:overlap val="-27" />
            <c:axId val="981700288" />
            <c:axId val="981700704" />
         </c:barChart>
         <c:catAx>
            <c:axId val="981700288" />
            <c:scaling>
               <c:orientation val="minMax" />
            </c:scaling>
            <c:delete val="0" />
            <c:axPos val="b" />
            <c:numFmt formatCode="General" sourceLinked="1" />
            <c:majorTickMark val="none" />
            <c:minorTickMark val="none" />
            <c:tickLblPos val="nextTo" />
            <c:spPr>
               <a:noFill />
               <a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">
                  <a:solidFill>
                     <a:schemeClr val="tx1">
                        <a:lumMod val="15000" />
                        <a:lumOff val="85000" />
                     </a:schemeClr>
                  </a:solidFill>
                  <a:round />
               </a:ln>
               <a:effectLst />
            </c:spPr>
            <c:txPr>
               <a:bodyPr rot="-60000000" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1" />
               <a:lstStyle />
               <a:p>
                  <a:pPr>
                     <a:defRPr sz="900" b="0" i="0" u="none" strike="noStrike" kern="1200" baseline="0">
                        <a:solidFill>
                           <a:schemeClr val="tx1">
                              <a:lumMod val="65000" />
                              <a:lumOff val="35000" />
                           </a:schemeClr>
                        </a:solidFill>
                        <a:latin typeface="+mn-lt" />
                        <a:ea typeface="+mn-ea" />
                        <a:cs typeface="+mn-cs" />
                     </a:defRPr>
                  </a:pPr>
                  <a:endParaRPr lang="ru-RU" />
               </a:p>
            </c:txPr>
            <c:crossAx val="981700704" />
            <c:crosses val="autoZero" />
            <c:auto val="1" />
            <c:lblAlgn val="ctr" />
            <c:lblOffset val="100" />
            <c:noMultiLvlLbl val="0" />
         </c:catAx>
         <c:valAx>
            <c:axId val="981700704" />
            <c:scaling>
               <c:orientation val="minMax" />
            </c:scaling>
            <c:delete val="0" />
            <c:axPos val="l" />
            <c:majorGridlines>
               <c:spPr>
                  <a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">
                     <a:solidFill>
                        <a:schemeClr val="tx1">
                           <a:lumMod val="15000" />
                           <a:lumOff val="85000" />
                        </a:schemeClr>
                     </a:solidFill>
                     <a:round />
                  </a:ln>
                  <a:effectLst />
               </c:spPr>
            </c:majorGridlines>
            <c:numFmt formatCode="#\ ##0.00\ &amp;quot;₽&amp;quot;" sourceLinked="1" />
            <c:majorTickMark val="none" />
            <c:minorTickMark val="none" />
            <c:tickLblPos val="nextTo" />
            <c:spPr>
               <a:noFill />
               <a:ln>
                  <a:noFill />
               </a:ln>
               <a:effectLst />
            </c:spPr>
            <c:txPr>
               <a:bodyPr rot="-60000000" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1" />
               <a:lstStyle />
               <a:p>
                  <a:pPr>
                     <a:defRPr sz="900" b="0" i="0" u="none" strike="noStrike" kern="1200" baseline="0">
                        <a:solidFill>
                           <a:schemeClr val="tx1">
                              <a:lumMod val="65000" />
                              <a:lumOff val="35000" />
                           </a:schemeClr>
                        </a:solidFill>
                        <a:latin typeface="+mn-lt" />
                        <a:ea typeface="+mn-ea" />
                        <a:cs typeface="+mn-cs" />
                     </a:defRPr>
                  </a:pPr>
                  <a:endParaRPr lang="ru-RU" />
               </a:p>
            </c:txPr>
            <c:crossAx val="981700288" />
            <c:crosses val="autoZero" />
            <c:crossBetween val="between" />
         </c:valAx>
         <c:spPr>
            <a:noFill />
            <a:ln>
               <a:noFill />
            </a:ln>
            <a:effectLst />
         </c:spPr>
      </c:plotArea>
      <c:plotVisOnly val="1" />
      <c:dispBlanksAs val="gap" />
      <c:showDLblsOverMax val="0" />
   </c:chart>
   <c:spPr>
      <a:solidFill>
         <a:schemeClr val="bg1" />
      </a:solidFill>
      <a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">
         <a:solidFill>
            <a:schemeClr val="tx1">
               <a:lumMod val="15000" />
               <a:lumOff val="85000" />
            </a:schemeClr>
         </a:solidFill>
         <a:round />
      </a:ln>
      <a:effectLst />
   </c:spPr>
   <c:txPr>
      <a:bodyPr />
      <a:lstStyle />
      <a:p>
         <a:pPr>
            <a:defRPr />
         </a:pPr>
         <a:endParaRPr lang="ru-RU" />
      </a:p>
   </c:txPr>
   <c:printSettings>
      <c:headerFooter />
      <c:pageMargins b="0.75" l="0.7" r="0.7" t="0.75" header="0.3" footer="0.3" />
      <c:pageSetup />
   </c:printSettings>
</c:chartSpace>''')

print l.legend

# Заголовок BarChart
for p in l.title.tx.rich.p:
    print p.text.t

for s in l.ser:
    print "lol 1: ", s.cat.strRef.f
    print "lol 2: " , s.val.numRef.f



# print l.type
# print l.style
# print l.grouping
# print l.overlap
# print l.title

# for s in l.ser:
#     # print s.xVal.numRef.numCache # Значения числовые
#     print s.xVal.numRef.f # Диапазон данных
#     print s.yVal.numRef.f

# self.idx = idx
# self.order = order
# self.tx = tx
# if spPr is None:
#     spPr = GraphicalProperties()
# self.spPr = spPr
# self.pictureOptions = pictureOptions
# self.dPt = dPt
# self.dLbls = dLbls
# self.trendline = trendline
# self.errBars = errBars
# self.cat = cat
# self.val = val
# self.invertIfNegative = invertIfNegative
# self.shape = shape
# self.xVal = xVal
# self.yVal = yVal
# self.bubbleSize = bubbleSize
# self.bubble3D = bubble3D
# if marker is None:
#     marker = Marker()
# self.marker = marker
# self.smooth = smooth
# self.explosion = explosion