# -*- coding: UTF-8 -*-
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout
# import xlsxwriter
# import xlwings as xw

reload(sys)
sys.setdefaultencoding('utf8')

# wb2 = load_workbook('lab1_template.xlsx')
# ws = wb2[wb2.get_sheet_names()[0]]

# print ws._charts
# print wb2.get_sheet_names()
# print type(ws.auto_filter.ref)
# # print ws.auto_filter
# first_cell = ws.auto_filter.ref.split(':')[0]
# cell_date_filter = ''
# cell_cost_filter = ''
# print ws[first_cell].row
# for Colfilter in ws.auto_filter.filterColumn:
#     if Colfilter.filters is not None:
#         cell_date_filter = ws[first_cell].column + str(Colfilter.colId)
#         print cell_date_filter
#         for Colfilter2 in Colfilter.filters.dateGroupItem:
#             print Colfilter2.year
#     print Colfilter.colId #относительно ws.auto_filter.ref
#     if Colfilter.customFilters is not None:
#         # print Colfilter.colId
#         for Colfilter1 in Colfilter.customFilters.customFilter:
#             print Colfilter1.operator+": "+Colfilter1.val
#             pass



def string_is_money_format(str):
    if '₽' in str:
        return True
    if '£' in str:
        return True
    if '$' in str:
        return True
    if '€' in str:
        return True
    return False


def range_is_money_rub_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not '₽' in cell.number_format:
                return {'status': False, 'message': 'Money rub format invalid'}
    return {'status': True, 'message': 'Money rub format valid'}

def range_is_money_dollar_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not '$' in cell.number_format:
                return {'status': False, 'message': 'Money dollar format invalid'}
    return {'status': True, 'message': 'Money dollar format valid'}

def range_is_formula_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if str(cell.value)[0] != '=':
                return False
    return True

def range_is_date_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not cell.is_date:
                return {'status': False, 'message': 'Dates invalid'}
    return {'status': True, 'message': 'Dates valid'}

def range_values_array(ws, range):
    arr = []
    rows = ws[range]
    for row in rows:
        for cell in row:
            arr.append(cell.value)
    return arr

def range_values_array_numeric(ws, range):
    arr = []
    rows = ws[range]
    for row in rows:
        for cell in row:
            arr.append(round(float(cell.value), 1))
    return arr

def calculate_correct_values(ws, range, employees, dollar_rate):
    data = {}
    data["names"] = employees
    try:
        rows = ws[range]
        values = []
        for row in rows:
            for cell in row:
                values.append(float(cell.value))
        data["salary"] = {"values": values}

        premium = []
        total = []
        income_tax = []
        amount_granted = []
        amount_granted_dollar = []

        premium_formula = []
        total_formula = []
        income_tax_formula = []
        amount_granted_formula = []
        amount_granted_dollar_formula = []



        for index, v in enumerate(data["salary"]["values"]):

            premium_val = v * 0.2
            premium.append(premium_val)
            premium_formula.append('=0,2*E'+str(index+5))

            total_val = premium_val + v
            total.append(total_val)
            total_formula.append('=F'+str(index+5)+'+E'+str(index+5))

            income_tax_val = round(total_val*0.13, 2)
            income_tax.append(income_tax_val)
            income_tax_formula.append('=G'+str(index+5)+'*0,13')

            amount_granted_val = (total_val - income_tax_val)
            amount_granted.append(amount_granted_val)
            amount_granted_formula.append('=G'+str(index+5)+'-H'+str(index+5))

            amount_granted_dollar_val = round(amount_granted_val/dollar_rate, 3)
            amount_granted_dollar.append(amount_granted_dollar_val)
            amount_granted_dollar_formula.append('=I'+str(index+5)+'/$C$14')

        data["premium"] = {"values": premium, "formula": premium_formula}
        data["total"] = {"values": total, "formula": total_formula}
        data["total_sum"] = { "values": sum(total), "formula": '=SUM(G5:G11)'}
        data["income_tax"] = {"values": income_tax, "formula": income_tax_formula}
        data["income_tax_sum"] = { "values": sum(income_tax), "formula": '=SUM(H5:H11)'}
        data["amount_granted"] = {"values": amount_granted, "formula": amount_granted_formula}
        data["amount_granted_sum"] = {"values": sum(amount_granted), "formula": '=SUM(I5:I11)'}
        data["amount_granted_dollar"] = {"values": amount_granted_dollar, "formula": amount_granted_dollar_formula}
        data["amount_granted_dollar_sum"] = {"values": sum(amount_granted_dollar), "formula": '=SUM(J5:J11)'}

        data["avg_salary"] = {"values": round(reduce(lambda x, y: x + y, amount_granted) / len(amount_granted), 2), "formula": '=AVERAGE(I5:I11)'}
        data["min_salary"] = {"values": round(min(amount_granted), 2), "formula": '=MIN(I5:I11)'}
        data["max_salary"] = {"values": round(max(amount_granted), 2), "formula": '=MAX(I5:I11)'}
        return data
    except:
        return False


def get_range_data(ws, range, ws_data_only):
    try:
        data = {}
        formula =  range_values_array(ws, range)
        data["formula"] = formula
        values = range_values_array_numeric(ws_data_only, range)
        data["values"]= values
        return data
    except:
        return False

def check_names(correct_data, student_data):
    if correct_data["names"] == student_data["names"]:
        return {'status': True, 'message': 'Names valid'}
    else:
        return {'status': False, 'message': 'Names invalid'}

def check_formats(student_ws):
    # Проверяем правильность форматирования дат поступления
    print "Формат дат поступления: ", range_is_date_format(student_ws, 'D5:D11')['status']

    # Проверяем правильность форматирования оклада
    print "Формат оклада: ", range_is_money_rub_format(student_ws, 'E5:E11')['status']

    # Проверяем правильность форматирования премии
    print "Формат премии: ", range_is_money_rub_format(student_ws, 'F5:F11')['status']

    # Проверяем правильность форматирования итого
    print "Формат итого: ", range_is_money_rub_format(student_ws, 'G5:G11')['status']

    # Проверяем правильность форматирования подоходного налога
    print "Формат подоходного налога: ", range_is_money_rub_format(student_ws, 'H5:H11')['status']

    # Проверяем правильность форматирования суммы к выдаче
    print "Формат суммы к выдаче: ", range_is_money_rub_format(student_ws, 'I5:I11')['status']

    # Проверяем правильность форматирования суммы к выдаче в долларах
    print "Формат суммы к выдаче в долларах: ", range_is_money_dollar_format(student_ws, 'J5:J11')['status']


def formulas_arrays_is_equal(f1, f2):
    clean_f1 = []
    clean_f2 = []
    for f in f1:
        clean_f1.append(f.replace(" ", "").lower().replace(".", ","))
    for f in f2:
        clean_f2.append(f.replace(" ", "").lower().replace(".", ","))

    return clean_f1 == clean_f2

def formulas_is_equal(f1, f2):
    if f1 and f2:
        f1 = f1.replace(" ", "").lower().replace(".", ",")
        f2 = f2.replace(" ", "").lower().replace(".", ",")

        return f1 == f2

    else: return False


def check_formulas(ws, ws_read_only, correct_data):

    # Проверяем Итого
    total = get_range_data(ws, 'G5:G11', ws_read_only)
    total_formulas_msg = "Итого посчитано: "
    if total and formulas_arrays_is_equal(total["formula"], correct_data["total"]["formula"]) and total["values"] == correct_data["total"]["values"]:
        total_formulas_msg += 'True'
    else:
        total_formulas_msg += 'False'
    print total_formulas_msg

    # Проверяем сумму к выдаче
    amount_granted = get_range_data(ws, 'I5:I11', ws_read_only)
    amount_granted_formulas_msg = "Суммы к выдаче посчитаны: "
    if amount_granted and formulas_arrays_is_equal(amount_granted["formula"], correct_data["amount_granted"]["formula"]) and amount_granted["values"] == correct_data["amount_granted"]["values"]:
        amount_granted_formulas_msg += 'True'
    else:
        amount_granted_formulas_msg += 'False'
    print amount_granted_formulas_msg

    # Проверяем сумму к выдаче в долларах
    amount_granted_dollar = get_range_data(ws, 'J5:J11', ws_read_only)
    amount_granted_dollar_formulas_msg = "Суммы к выдаче в долларах посчитаны: "

    if amount_granted_dollar and formulas_arrays_is_equal(amount_granted_dollar["formula"], correct_data["amount_granted_dollar"]["formula"]) and amount_granted_dollar["values"] == correct_data["amount_granted_dollar"]["values"]:
        amount_granted_dollar_formulas_msg += 'True'
    else:
        amount_granted_dollar_formulas_msg += 'False'
    print amount_granted_dollar_formulas_msg


    # TODO: доделать проверку формул
    # Проверяем премии

    premium = get_range_data(ws, 'F5:F11', ws_read_only)

    premium_formulas_msg = "Премии посчитаны: "
    if premium and formulas_arrays_is_equal(premium["formula"], correct_data["premium"]["formula"]) and premium["values"] == correct_data["premium"]["values"]:
        premium_formulas_msg += 'True'
    else:
        premium_formulas_msg += 'False'
    print premium_formulas_msg

    # Проверяем подоходный налог
    income_tax = get_range_data(ws, 'H5:H11', ws_read_only)
    income_tax_formulas_msg = "Подоходный налог посчитан: "
    if income_tax and formulas_arrays_is_equal(income_tax["formula"], correct_data["income_tax"]["formula"]) and income_tax["values"] == correct_data["income_tax"]["values"]:
        income_tax_formulas_msg += 'True'
    else:
        income_tax_formulas_msg += 'False'
    print income_tax_formulas_msg

def check_functions(ws, ws_read_only, correct_data):

    # Проверяем сумму Итого
    total_sum_func_msg = 'Сумма Итого: '
    total_sum_cell = 'G12'
    if formulas_is_equal(ws[total_sum_cell].value, correct_data["total_sum"]["formula"]) and round(ws_read_only[total_sum_cell].value, 1) == round(correct_data["total_sum"]["values"], 1):
        total_sum_func_msg += 'True'
    else:
        total_sum_func_msg += 'False'
    print total_sum_func_msg

    # Проверяем сумму подоходных налогов
    income_tax_sum_func_msg = 'Сумма подоходного налога: '
    income_tax_sum_cell = 'H12'
    if formulas_is_equal(ws[income_tax_sum_cell].value, correct_data["income_tax_sum"]["formula"]) and round(ws_read_only[income_tax_sum_cell].value, 1) == round(correct_data["income_tax_sum"]["values"], 1):
        income_tax_sum_func_msg += 'True'
    else:
        income_tax_sum_func_msg += 'False'
    print income_tax_sum_func_msg

    # Проверяем сумму сумм к выдаче
    amount_granted_sum_func_msg = 'Сумма к выдаче: '
    amount_granted_sum_cell = 'I12'
    if formulas_is_equal(ws[amount_granted_sum_cell].value, correct_data["amount_granted_sum"]["formula"]) and round(ws_read_only[amount_granted_sum_cell].value, 1) == round(correct_data["amount_granted_sum"]["values"], 1):
        amount_granted_sum_func_msg += 'True'
    else:
        amount_granted_sum_func_msg += 'False'
    print amount_granted_sum_func_msg


    # Проверяем сумму сумм к выдаче в долларах
    amount_granted_dollar_sum_func_msg = 'Сумма к выдаче в долларах: '
    amount_granted_dollar_sum_cell = 'J12'
    if formulas_is_equal(ws[amount_granted_dollar_sum_cell].value, correct_data["amount_granted_dollar_sum"]["formula"]) and round(ws_read_only[amount_granted_dollar_sum_cell].value, 1) == round(correct_data["amount_granted_dollar_sum"]["values"], 1):
        amount_granted_dollar_sum_func_msg += 'True'
    else:
        amount_granted_dollar_sum_func_msg += 'False'
    print amount_granted_dollar_sum_func_msg

    # Проверяем среднее значение суммы к выдаче
    avg_salary_func_msg = 'Среднее значение зарплаты: '
    avg_salary_sum_cell = 'C15'
    if formulas_is_equal(ws[avg_salary_sum_cell].value, correct_data["avg_salary"]["formula"]) and round(ws_read_only[avg_salary_sum_cell].value, 1) == round(correct_data["avg_salary"]["values"], 1):
        avg_salary_func_msg += 'True'
    else:
        avg_salary_func_msg += 'False'
    print avg_salary_func_msg

    # Проверяем максимальное значение суммы к выдаче
    max_salary_func_msg = 'Максимальное значение зарплаты: '
    max_salary_sum_cell = 'C16'
    if formulas_is_equal(ws[max_salary_sum_cell].value, correct_data["max_salary"]["formula"]) and round(ws_read_only[max_salary_sum_cell].value, 1) == round(correct_data["max_salary"]["values"], 1):
        max_salary_func_msg += 'True'
    else:
        max_salary_func_msg += 'False'
    print max_salary_func_msg

    # Проверяем минимальное значение суммы к выдаче
    min_salary_func_msg = 'Минимальное значение зарплаты: '
    min_salary_sum_cell = 'C17'
    if formulas_is_equal(ws[min_salary_sum_cell].value, correct_data["min_salary"]["formula"]) and round(ws_read_only[min_salary_sum_cell].value, 1) == round(correct_data["min_salary"]["values"], 1):
        min_salary_func_msg += 'True'
    else:
        min_salary_func_msg += 'False'
    print min_salary_func_msg

def check_ws_have_rule(ws, cells_range, operator, formula_value):
    for rule in ws.conditional_formatting.cf_rules.items():
        if cells_range == rule[0]:
            if operator == rule[1][0].operator:
                if int(rule[1][0].formula[0]) == formula_value:
                    print 'Условное форматирование выполнено: True'
    print 'Условное форматирование выполнено: False'



def check_answer(student_wb, student_wb_data_only, employees, dollar_rate):
    student_ws = student_wb[student_wb.get_sheet_names()[0]]
    ws_read_only = student_wb_data_only[student_wb_data_only.get_sheet_names()[0]]

    start_row = 5
    end_row = start_row + len(employees) - 1
    range_salary = 'E5:E11'

    correct_values_data = calculate_correct_values(student_ws, range_salary, sorted(employees), dollar_rate)

    if correct_values_data:
        print 'Столбец окладов заполенен'
        # Проверяем правильность ФИО (в т.ч. сортировку)
        check_formats(student_ws)

        # Проверем правильность формул и значений
        check_formulas(student_ws, ws_read_only, correct_values_data)

        # Проверяем правильность функций
        check_functions(student_ws, ws_read_only, correct_values_data)

        # Проверем правильность условного форматирования
        check_ws_have_rule(student_ws, 'I5:I11', 'lessThan', 5000)
    else:
        message = 'Неверно заполнен столбец "Оклад"'