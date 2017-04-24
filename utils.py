# -*- coding: UTF-8 -*-
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout

reload(sys)
sys.setdefaultencoding('utf8')

def range_is_date_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not cell.is_date:
                return {'status': False, 'message': 'Dates invalid'}
    return {'status': True, 'message': 'Dates valid'}

def range_is_money_rub_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not '₽' in cell.number_format:
                return {'status': False, 'message': 'Money rub format invalid'}
    return {'status': True, 'message': 'Money rub format valid'}

def range_is_money_dollar_format(ws, range):
    rows = ws[range]
    for row in rows:
        for cell in row:
            if not '$' in cell.number_format:
                return {'status': False, 'message': 'Money dollar format invalid'}
    return {'status': True, 'message': 'Money dollar format valid'}

def formulas_is_equal(f1, f2):
    if f1 and f2:
        f1 = f1.replace(" ", "").lower().replace(".", ",")
        f2 = f2.replace(" ", "").lower().replace(".", ",")

        return f1 == f2

    else: return False