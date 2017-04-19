# -*- coding: UTF-8 -*-
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.reader import reader
from openpyxl.chart.layout import Layout, ManualLayout
# import xlsxwriter
# import xlwings as xw

reload(sys)
sys.setdefaultencoding('utf8')

wb2 = load_workbook('data/rabota 8.xlsx')
ws = wb2[wb2.get_sheet_names()[0]]


# print wb2.get_sheet_names()
# print type(ws.auto_filter.ref)
# # print ws.auto_filter
# first_cell = ws.auto_filter.ref.split(':')[0]
# cell_date_filter = ''
# cell_cost_filter = ''
# print ws[first_cell].row
# for Colfilter in ws.auto_filter.filterColumn:
#     if Colfilter.filters is not None:
#         cell_date_filter = ws[first_cell].column + str(Colfilter.colId)
#         print cell_date_filter
#         for Colfilter2 in Colfilter.filters.dateGroupItem:
#             print Colfilter2.year
#     print Colfilter.colId #относительно ws.auto_filter.ref
#     if Colfilter.customFilters is not None:
#         # print Colfilter.colId
#         for Colfilter1 in Colfilter.customFilters.customFilter:
#             print Colfilter1.operator+": "+Colfilter1.val
#             pass


# interesCell = 'G4'
#
# print "D7: ", ws[interesCell].value
# print "D7: ", ws[interesCell].coordinate
# print "D7: ", ws[interesCell].column
# print "D7: ", ws[interesCell].base_date
# print "D7: ", ws[interesCell].guess_types
# print "D7: ", ws[interesCell].internal_value
# print "D7: ", ws[interesCell].is_date
# print "D7: ", ws[interesCell].number_format


l = reader('''<?xml version="1.0" encoding="UTF-8"?>
<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:c16r2="http://schemas.microsoft.com/office/drawing/2015/06/chart" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
   <c:date1904 val="0" />
   <c:lang val="ru-RU" />
   <c:roundedCorners val="0" />
   <mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006">
      <mc:Choice xmlns:c14="http://schemas.microsoft.com/office/drawing/2007/8/2/chart" Requires="c14">
         <c14:style val="102" />
      </mc:Choice>
      <mc:Fallback>
         <c:style val="2" />
      </mc:Fallback>
   </mc:AlternateContent>
   <c:chart>
      <c:title>
         <c:tx>
            <c:rich>
               <a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1" />
               <a:lstStyle />
               <a:p>
                  <a:pPr>
                     <a:defRPr sz="1400" b="0" i="0" u="none" strike="noStrike" kern="1200" spc="0" baseline="0">
                        <a:solidFill>
                           <a:schemeClr val="tx1">
                              <a:lumMod val="65000" />
                              <a:lumOff val="35000" />
                           </a:schemeClr>
                        </a:solidFill>
                        <a:latin typeface="+mn-lt" />
                        <a:ea typeface="+mn-ea" />
                        <a:cs typeface="+mn-cs" />
                     </a:defRPr>
                  </a:pPr>
                  <a:r>
                     <a:rPr lang="ru-RU" />
                     <a:t>КРУГ диаграммы</a:t>
                  </a:r>
               </a:p>
            </c:rich>
         </c:tx>
         <c:layout />
         <c:overlay val="0" />
         <c:spPr>
            <a:noFill />
            <a:ln>
               <a:noFill />
            </a:ln>
            <a:effectLst />
         </c:spPr>
         <c:txPr>
            <a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" anchor="ctr" anchorCtr="1" />
            <a:lstStyle />
            <a:p>
               <a:pPr>
                  <a:defRPr sz="1400" b="0" i="0" u="none" strike="noStrike" kern="1200" spc="0" baseline="0">
                     <a:solidFill>
                        <a:schemeClr val="tx1">
                           <a:lumMod val="65000" />
                           <a:lumOff val="35000" />
                        </a:schemeClr>
                     </a:solidFill>
                     <a:latin typeface="+mn-lt" />
                     <a:ea typeface="+mn-ea" />
                     <a:cs typeface="+mn-cs" />
                  </a:defRPr>
               </a:pPr>
               <a:endParaRPr lang="ru-RU" />
            </a:p>
         </c:txPr>
      </c:title>
      <c:autoTitleDeleted val="0" />
      <c:plotArea>
         <c:layout />
         <c:pieChart>
            <c:varyColors val="1" />
            <c:ser>
               <c:idx val="0" />
               <c:order val="0" />
               <c:dPt>
                  <c:idx val="0" />
                  <c:bubble3D val="0" />
                  <c:spPr>
                     <a:solidFill>
                        <a:schemeClr val="accent1" />
                     </a:solidFill>
                     <a:ln w="19050">
                        <a:solidFill>
                           <a:schemeClr val="lt1" />
                        </a:solidFill>
                     </a:ln>
                     <a:effectLst />
                  </c:spPr>
               </c:dPt>
               <c:dPt>
                  <c:idx val="1" />
                  <c:bubble3D val="0" />
                  <c:spPr>
                     <a:solidFill>
                        <a:schemeClr val="accent2" />
                     </a:solidFill>
                     <a:ln w="19050">
                        <a:solidFill>
                           <a:schemeClr val="lt1" />
                        </a:solidFill>
                     </a:ln>
                     <a:effectLst />
                  </c:spPr>
               </c:dPt>
               <c:dLbls>
                  <c:dLbl>
                     <c:idx val="0" />
                     <c:layout>
                        <c:manualLayout>
                           <c:x val="3.2998906386701665E-2" />
                           <c:y val="1.9644575678040246E-2" />
                        </c:manualLayout>
                     </c:layout>
                     <c:showLegendKey val="0" />
                     <c:showVal val="1" />
                     <c:showCatName val="0" />
                     <c:showSerName val="0" />
                     <c:showPercent val="0" />
                     <c:showBubbleSize val="0" />
                  </c:dLbl>
                  <c:dLbl>
                     <c:idx val="1" />
                     <c:layout>
                        <c:manualLayout>
                           <c:x val="-6.9050087489063869E-2" />
                           <c:y val="-9.6518664333624965E-2" />
                        </c:manualLayout>
                     </c:layout>
                     <c:showLegendKey val="0" />
                     <c:showVal val="1" />
                     <c:showCatName val="0" />
                     <c:showSerName val="0" />
                     <c:showPercent val="0" />
                     <c:showBubbleSize val="0" />
                  </c:dLbl>
                  <c:spPr>
                     <a:noFill />
                     <a:ln>
                        <a:noFill />
                     </a:ln>
                     <a:effectLst />
                  </c:spPr>
                  <c:txPr>
                     <a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" vert="horz" wrap="square" lIns="38100" tIns="19050" rIns="38100" bIns="19050" anchor="ctr" anchorCtr="1">
                        <a:spAutoFit />
                     </a:bodyPr>
                     <a:lstStyle />
                     <a:p>
                        <a:pPr>
                           <a:defRPr sz="900" b="0" i="0" u="none" strike="noStrike" kern="1200" baseline="0">
                              <a:solidFill>
                                 <a:schemeClr val="tx1">
                                    <a:lumMod val="75000" />
                                    <a:lumOff val="25000" />
                                 </a:schemeClr>
                              </a:solidFill>
                              <a:latin typeface="+mn-lt" />
                              <a:ea typeface="+mn-ea" />
                              <a:cs typeface="+mn-cs" />
                           </a:defRPr>
                        </a:pPr>
                        <a:endParaRPr lang="ru-RU" />
                     </a:p>
                  </c:txPr>
                  <c:showLegendKey val="0" />
                  <c:showVal val="1" />
                  <c:showCatName val="0" />
                  <c:showSerName val="0" />
                  <c:showPercent val="0" />
                  <c:showBubbleSize val="0" />
                  <c:showLeaderLines val="1" />
                  <c:leaderLines>
                     <c:spPr>
                        <a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">
                           <a:solidFill>
                              <a:schemeClr val="tx1">
                                 <a:lumMod val="35000" />
                                 <a:lumOff val="65000" />
                              </a:schemeClr>
                           </a:solidFill>
                           <a:round />
                        </a:ln>
                        <a:effectLst />
                     </c:spPr>
                  </c:leaderLines>
               </c:dLbls>
               <c:val>
                  <c:numRef>
                     <c:f>Sheet!$H$12:$I$12</c:f>
                     <c:numCache>
                        <c:formatCode>#\ ##0.00\ "₽"</c:formatCode>
                        <c:ptCount val="2" />
                        <c:pt idx="0">
                           <c:v>15607.8</c:v>
                        </c:pt>
                        <c:pt idx="1">
                           <c:v>104452.2</c:v>
                        </c:pt>
                     </c:numCache>
                  </c:numRef>
               </c:val>
            </c:ser>
            <c:dLbls>
               <c:showLegendKey val="0" />
               <c:showVal val="0" />
               <c:showCatName val="0" />
               <c:showSerName val="0" />
               <c:showPercent val="0" />
               <c:showBubbleSize val="0" />
               <c:showLeaderLines val="1" />
            </c:dLbls>
            <c:firstSliceAng val="0" />
         </c:pieChart>
         <c:spPr>
            <a:noFill />
            <a:ln>
               <a:noFill />
            </a:ln>
            <a:effectLst />
         </c:spPr>
      </c:plotArea>
      <c:plotVisOnly val="1" />
      <c:dispBlanksAs val="gap" />
      <c:showDLblsOverMax val="0" />
   </c:chart>
   <c:spPr>
      <a:solidFill>
         <a:schemeClr val="bg1" />
      </a:solidFill>
      <a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">
         <a:solidFill>
            <a:schemeClr val="tx1">
               <a:lumMod val="15000" />
               <a:lumOff val="85000" />
            </a:schemeClr>
         </a:solidFill>
         <a:round />
      </a:ln>
      <a:effectLst />
   </c:spPr>
   <c:txPr>
      <a:bodyPr />
      <a:lstStyle />
      <a:p>
         <a:pPr>
            <a:defRPr />
         </a:pPr>
         <a:endParaRPr lang="ru-RU" />
      </a:p>
   </c:txPr>
   <c:printSettings>
      <c:headerFooter />
      <c:pageMargins b="0.75" l="0.7" r="0.7" t="0.75" header="0.3" footer="0.3" />
      <c:pageSetup />
   </c:printSettings>
</c:chartSpace>''')

print l.legend

# Заголовок BarChart
# for p in l.title.tx.rich.p:
#     print p.text.t

# for s in l.ser:
#     print "lol 1: ", s.cat.strRef.f
#     print "lol 2: " , s.val.numRef.f



