# -*- coding: UTF-8 -*-
import sys
import random
import datetime
import json

from openpyxl import Workbook, load_workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

from lab_3_create_template import lab_3_create_template
from lab_3_check_answer import lab_3_check_answer

reload(sys)
sys.setdefaultencoding('utf8')


data = [
    ["Комбайн", "19.07.2017", 100, 7800.00],
    ["Миксер", "30.05.2017", 38, 3000.00],
    ["Микровоновка", "23.08.2017", 38, 4500.00],
    ["Пылесос", "17.03.2017", 25, 3000.00],
    ["Холодильник", "03.05.2016", 56, 25000.00],
    ["Пылесос", "03.08.2017", 6, 1500.00],
    ["Телевизор", "02.03.2014", 50, 6000.00],
    ["Телевизор", "16.02.2016", 19, 12000.00],
    ["Телевизор", "13.09.2017", 32, 4500.00],
    ["Утюг", "12.07.2016", 70, 2000.00],
    ["Утюг", "20.08.2016", 15, 1000.00],
    ["Утюг", "02.08.2017", 20, 2900.00],
    ["Чайник", "15.03.2017", 25, 1540.00],
    ["Чайник", "27.07.2016", 102, 1200.00],
    ["Чайник", "04.08.2016", 45, 500.00],
]

template_wb = Workbook()
template_ws = template_wb.active

template_ws = lab_3_create_template(template_ws)
template_wb.save('lab3_template.xlsx')

correct_wb = load_workbook('lab3_correct.xlsx')
correct_wb_data_only =  load_workbook('lab3_correct.xlsx', data_only=True)

# Проверка
student_wb =  load_workbook('lab3_correct.xlsx')
student_wb_data_only =  load_workbook('lab3_correct.xlsx', data_only=True)

result = lab_3_check_answer(correct_wb, correct_wb_data_only, student_wb, student_wb_data_only)

print json.dumps(result)





